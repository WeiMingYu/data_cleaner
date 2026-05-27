# layer3_export/exporter.py
# 匯出主邏輯
# 把比對結果寫成 Excel
# 產出：
#   Sheet 1：更新後的主檔（原有資料 + 補齊）
#   Sheet 2：新增資料（來源有、主檔沒有）
#   Sheet 3：待人工確認（比對不到的）

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import (
    OUTPUT_FILE, MASTER_FILE,
    UNMATCHED_SHEET_NAME, NEW_RECORDS_SHEET_NAME,
)
from master_schema import MASTER_COLUMNS


class Exporter:
    """
    比對結果匯出器。
    把三種資料寫成多 sheet 的 Excel 檔。
    """

    # 顏色定義（openpyxl PatternFill 用 ARGB 格式）
    COLOR_HEADER   = "FF2B5797"   # 深藍（標題列）
    COLOR_NEW      = "FFFFC000"   # 橙黃（新增資料標記）
    COLOR_UNMATCH  = "FFFF0000"   # 紅（待確認標記）
    COLOR_FILLED   = "FFE2EFDA"   # 淺綠（已補齊欄位，可選）

    def __init__(
        self,
        updated_master_df: pd.DataFrame,
        new_df: pd.DataFrame,
        unmatched_df: pd.DataFrame,
        output_file: str = OUTPUT_FILE,
    ):
        self.updated_master_df = updated_master_df
        self.new_df = new_df
        self.unmatched_df = unmatched_df
        self.output_file = output_file

    # ===== 1. 主輸出流程 =====

    def export(self):
        """把三個 DataFrame 寫成 Excel，回傳輸出路徑"""
        print(f"\n[Layer3] 開始匯出：{self.output_file}")

        # 確保輸出目錄存在
        output_dir = os.path.dirname(self.output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)

        # 用 ExcelWriter 寫多個 sheet
        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:

            # --- Sheet 1：主檔（補齊後）---
            self.updated_master_df.to_excel(
                writer, sheet_name="主檔", index=False
            )

            # --- Sheet 2：新增資料 ---
            if len(self.new_df) > 0:
                self.new_df.to_excel(
                    writer, sheet_name=NEW_RECORDS_SHEET_NAME, index=False
                )
            else:
                # 即使沒有新資料也建一個空 sheet，方便辨識
                pd.DataFrame(columns=MASTER_COLUMNS).to_excel(
                    writer, sheet_name=NEW_RECORDS_SHEET_NAME, index=False
                )

            # --- Sheet 3：待人工確認 ---
            if len(self.unmatched_df) > 0:
                self.unmatched_df.to_excel(
                    writer, sheet_name=UNMATCHED_SHEET_NAME, index=False
                )
            else:
                pd.DataFrame(columns=MASTER_COLUMNS + ["比對備註"]).to_excel(
                    writer, sheet_name=UNMATCHED_SHEET_NAME, index=False
                )

        # 套用樣式
        self._apply_styles()

        print(f"  主檔：{len(self.updated_master_df)} 筆")
        print(f"  新增資料：{len(self.new_df)} 筆")
        print(f"  待人工確認：{len(self.unmatched_df)} 筆")
        print(f"  輸出完成：{self.output_file}")
        return self.output_file

    # ===== 2. 套用樣式 =====

    def _apply_styles(self):
        """開啟產出的 xlsx，套用標題色、欄寬等樣式"""
        wb = load_workbook(self.output_file)

        for ws in wb.worksheets:
            self._style_sheet(ws)

        wb.save(self.output_file)

    def _style_sheet(self, ws):
        """對單一 sheet 套用樣式"""
        # --- 標題列樣式 ---
        header_fill = PatternFill("solid", fgColor=self.COLOR_HEADER)
        header_font = Font(bold=True, color="FFFFFFFF", name="微軟正黑體")
        header_align = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # --- 凍結標題列 ---
        ws.freeze_panes = "A2"

        # --- 自動欄寬（最大 50 個字元）---
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value else 0
                    # 中文字算兩個寬度
                    chinese_chars = sum(1 for c in str(cell.value or "") if '\u4e00' <= c <= '\u9fff')
                    cell_len = cell_len + chinese_chars
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            # 最小 8，最大 40
            ws.column_dimensions[col_letter].width = max(8, min(40, max_len + 2))

        # --- 資料列字型 ---
        data_font = Font(name="微軟正黑體", size=10)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = data_font

    # ===== 3. 印出摘要 =====

    def print_summary(self):
        """印出輸出摘要"""
        print("\n" + "=" * 50)
        print("[Layer3] 匯出摘要")
        print("=" * 50)
        print(f"  輸出檔案 ：{self.output_file}")
        print(f"  主檔筆數 ：{len(self.updated_master_df)}")
        print(f"  新增筆數 ：{len(self.new_df)}")
        print(f"  待確認筆數：{len(self.unmatched_df)}")
        print("=" * 50)
